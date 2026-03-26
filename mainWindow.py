import os
import sys
from pathlib import Path
from tkinter.filedialog import dialogstates

import PyQt5
import PyQt5.QtCore as QtCore
from PyQt5 import Qt
from PyQt5.QtWidgets import QHBoxLayout, QVBoxLayout, QWidget, QTextEdit, QMainWindow, QApplication, QAction, \
    QDesktopWidget, QLabel, QLineEdit, QPushButton, QFileDialog, QStackedWidget, QFormLayout
import pandas as pd
import openpyxl
from openpyxl import load_workbook

print("pyqt5:v"+QtCore.QT_VERSION_STR)
print(QtCore.QT_VERSION_STR)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.error = '<font color="red">{}</font>'
        self.warning = '<font color="orange">{}</font>'
        self.valid = '<font color="green">{}</font>'

        #设置菜单栏
        self.ex = None
        self.set_bar()

        #功能模块按键
        self.btn_mod1 = QPushButton("数据统计")
        self.btn_mod2 = QPushButton("财务报表")
        self.btn_mod3 = QPushButton("工资核算")
        self.btn_mod4 = QPushButton("差旅报销")

        self.btn_mod1 = QPushButton("数据统计",self)
        self.btn_mod2 = QPushButton("财务报表",self)
        self.btn_mod3 = QPushButton("工资核算",self)
        self.btn_mod4 = QPushButton("差旅报销", self)

        self.btn_mod1.setFixedSize(80,40)
        self.btn_mod2.setFixedSize(80,40)
        self.btn_mod3.setFixedSize(80,40)
        self.btn_mod4.setFixedSize(80,40)

        self.btn_mod1.setStyleSheet("background-color: rgb(190,226,224); color: black; border-radius: 0px;")
        self.btn_mod2.setStyleSheet("background-color: rgb(255,255,255); color: black; border-radius: 0px;")
        self.btn_mod3.setStyleSheet("background-color: rgb(255,255,255); color: black; border-radius: 0px;")
        self.btn_mod4.setStyleSheet("background-color: rgb(255,255,255); color: black; border-radius: 0px;")

        self.btn_mod1.clicked.connect(self.btn_mod1_clicked)
        self.btn_mod2.clicked.connect(self.btn_mod2_clicked)
        self.btn_mod3.clicked.connect(self.btn_mod3_clicked)
        self.btn_mod4.clicked.connect(self.btn_mod4_clicked)

        #放置一个central_widget
        central_widget = QWidget(self)
        self.setCentralWidget(central_widget)

        #设置水平布局及标题
        main_layout = QVBoxLayout()
        layout_1 = QHBoxLayout()
        layout_2 = QHBoxLayout()

        #堆栈窗口
        self.stack1 = QWidget(self)
        self.stack2 = QWidget(self)
        self.stack3 = QWidget(self)
        self.stack4 = QWidget(self)

        #stack1组件
        self.btn = None
        self.path = None
        self.line_edit_path = None
        self.push_button_start = None
        self.text_edit = None
        self.stack1_ui()

        # stack2组件
        self.stack2_ui()
        # stack3组件
        self.stack3_ui()
        # stack4组件
        self.stack4_ui()

        self.Stack = QStackedWidget (self)
        self.Stack.addWidget (self.stack1)
        self.Stack.addWidget (self.stack2)
        self.Stack.addWidget (self.stack3)
        self.Stack.addWidget (self.stack4)
        self.Stack.setStyleSheet("background-color: rgb(190,226,224);")

        #设置窗体的宽高和标题
        self.resize(1680,960)
        self.setWindowTitle("ChuChu小助手 V1.0")

        #将左侧和右侧布局添加到主水平布局中
        layout_1.addWidget(self.btn_mod1)
        layout_1.addWidget(self.btn_mod2)
        layout_1.addWidget(self.btn_mod3)
        layout_1.addWidget(self.btn_mod4)
        layout_1.setAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignTop)
        layout_1.setContentsMargins(0,0,0,0)         #控件到layout之间的距离
        layout_1.setSpacing(0)                        #控件之间的距离

        layout_2.addWidget(self.Stack)

        main_layout.addLayout(layout_1)
        main_layout.addLayout(layout_2)
        main_layout.setContentsMargins(0, 0, 0, 0)  # 控件到layout之间的距离
        main_layout.setSpacing(0)  # 控件之间的距离

        central_widget.setLayout(main_layout)
        #设置窗口的主布局
        self.setCentralWidget(central_widget)

        # 设置信息
        temp_path = QtCore.QDir(self.path)
        self.set_path = temp_path.absolutePath()+'/setting/'
        if not os.path.isdir(self.set_path):
            os.mkdir(self.set_path)
            print("set文件不存在")
            self.text_edit.append(self.error.format("请先在setting目录下添加设置文件"))
        else:
            set_file = Path(self.set_path + 'setting.xlsx')
            if set_file.is_file():
                try:
                    self.set_data = pd.read_excel(set_file, dtype=str)  # 以字符串形式打开并读取excel表格
                    self.set_data = pd.DataFrame(self.set_data)
                    print(self.set_data)
                    print('set文件存在')
                    self.text_edit.append(self.valid.format("设置文件读取正常"))
                except FileNotFoundError as e:
                    print(f"设置文件打开失败: {e}")
                    self.text_edit.append(self.error.format(f"设置文件打开失败: {e}"))
            else:
                print("set文件不存在")
                self.text_edit.append(self.error.format("设置文件不存在"))

        #目标写入文件
        self.des_sheet = None

        #调用居中函数
        self.show()
        self.center()

    def btn_mod1_clicked(self):
        self.Stack.setCurrentIndex(0)
        self.btn_mod1.setStyleSheet("background-color: rgb(190,226,224); color: black; border-radius: 0px;")
        self.btn_mod2.setStyleSheet("background-color: rgb(255,255,255);")
        self.btn_mod3.setStyleSheet("background-color: rgb(255,255,255);")
        self.btn_mod4.setStyleSheet("background-color: rgb(255,255,255);")

    def btn_mod2_clicked(self):
        self.Stack.setCurrentIndex(1)
        self.btn_mod1.setStyleSheet("background-color: rgb(255,255,255);")
        self.btn_mod2.setStyleSheet("background-color: rgb(190,226,224); color: black; border-radius: 0px;")
        self.btn_mod3.setStyleSheet("background-color: rgb(255,255,255);")
        self.btn_mod4.setStyleSheet("background-color: rgb(255,255,255);")

    def btn_mod3_clicked(self):
        self.Stack.setCurrentIndex(2)
        self.btn_mod1.setStyleSheet("background-color: rgb(255,255,255);")
        self.btn_mod2.setStyleSheet("background-color: rgb(255,255,255);")
        self.btn_mod3.setStyleSheet("background-color: rgb(190,226,224); color: black; border-radius: 0px;")
        self.btn_mod4.setStyleSheet("background-color: rgb(255,255,255);")

    def btn_mod4_clicked(self):
        self.Stack.setCurrentIndex(3)
        self.btn_mod1.setStyleSheet("background-color: rgb(255,255,255);")
        self.btn_mod2.setStyleSheet("background-color: rgb(255,255,255);")
        self.btn_mod3.setStyleSheet("background-color: rgb(255,255,255);")
        self.btn_mod4.setStyleSheet("background-color: rgb(190,226,224); color: black; border-radius: 0px;")

    def stack1_ui(self):
        print("stack1_ui")
        #实例化按键
        self.btn = QPushButton("打开需要统计的目录")
        self.btn.setStyleSheet("background-color: rgb(255,255,255); color: black;")
        self.btn.setFixedHeight(40)
        self.btn.clicked.connect(self.get_dir)

        #文本框设定
        self.path = QtCore.QDir.currentPath()
        self.line_edit_path = QLineEdit(str(self.path))
        self.line_edit_path.setFixedHeight(40)
        self.line_edit_path.setStyleSheet("QLineEdit { background-color: white; }")

        #开始按键
        self.push_button_start = QPushButton('开始')
        self.push_button_start.setStyleSheet("background-color: rgb(255,255,255); color: black;")
        self.push_button_start.setFixedSize(100, 40)
        self.push_button_start.clicked.connect(self.start_button)

        #实例化textEdit并加入布局
        self.text_edit = QTextEdit()
        self.text_edit.setStyleSheet("QTextEdit { background-color: white; }")

        stack1_main_layout = QVBoxLayout()
        stack1_layout_1 = QHBoxLayout()
        stack1_layout_2 = QHBoxLayout()

        #并加入布局
        stack1_layout_1.addWidget(self.btn)
        stack1_layout_1.addWidget(self.line_edit_path)
        stack1_layout_1.addWidget(self.push_button_start)
        stack1_layout_2.addWidget(self.text_edit)
        stack1_main_layout.addLayout(stack1_layout_1)
        stack1_main_layout.addLayout(stack1_layout_2)
        self.stack1.setLayout(stack1_main_layout)

    def stack2_ui(self):
        layout = QFormLayout()
        layout1 = QHBoxLayout()
        layout1.addWidget(QLabel("敬请期待1"))
        layout.addRow(layout1)
        self.stack2.setLayout(layout)

    def stack3_ui(self):
        layout = QFormLayout()
        layout1 = QHBoxLayout()
        layout1.addWidget(QLabel("敬请期待2"))
        layout.addRow(layout1)
        self.stack3.setLayout(layout)

    def stack4_ui(self):
        layout = QFormLayout()
        layout1 = QHBoxLayout()
        layout1.addWidget(QLabel("敬请期待3"))
        layout.addRow(layout1)
        self.stack4.setLayout(layout)

    def get_dir(self):
        dialog = QFileDialog()
        dialog.options = QFileDialog.Options()
        dialog.options |= QFileDialog.ShowDirsOnly
        folder_path = QFileDialog.getExistingDirectory(self, "选择文件夹", options=dialog.options)

        if folder_path:
            print(f"选择的文件夹：{folder_path}")
            self.path = folder_path
            self.line_edit_path.setText(folder_path)

    def open_set_file(self):
        dialog = QFileDialog(self, "打开设置文件")
        dialog.setDirectory(self.set_path)
        dialog.setFileMode(QFileDialog.AnyFile)
        dialog.setOption(QFileDialog.ReadOnly)

        if dialog.exec_():
            select_set_file = dialog.selectedFiles()[0]
            os.startfile(select_set_file)
            print(f"选择设置文件：{select_set_file}")

    def set_bar(self):
        # 实例化主窗口的QMenuBar对象
        bar = self.menuBar()

        # 向菜单栏中添加“文件”
        file = bar.addMenu('文件')
        # 向QMenu小控件中添加按钮，子菜单
        file.addAction('打开文件')
        # 定义响应小控件按钮，并设置快捷键关联到操作按钮，添加到父菜单下
        save = QAction('保存文件', self)
        save.setShortcut('Ctrl+S')
        file.addAction(save)
        # 父菜单下
        quit_ = QAction('退出', self)
        file.addAction(quit_)

        # 向菜单栏中添加“设置”
        menu_set = bar.addMenu('设置')
        resset_file = QAction('配置文件', self)
        menu_set.addAction(resset_file)
        menu_set.triggered[QAction].connect(self.open_set_file)

    def center(self):
        # 获取屏幕的大小
        screen = QDesktopWidget().screenGeometry()
        # 获取窗口的大小
        size = self.geometry()
        # 将窗口移动到屏幕中央
        self.move(int((screen.width() - size.width()) / 2), int((screen.height() - size.height()) / 2))

    @staticmethod
    def company_process(self, full_path, company_data):
        paths = os.walk(full_path)
        for path, dir_lst, file_lst in paths:
            for file_name in file_lst:
                file_path = os.path.join(path, file_name)
                file_name = os.path.basename(file_path)
                file = Path(str(file_path))
                print("文件:" + file_name)

                if file.is_file():
                    for index_row, row in company_data.iterrows():  # 遍历过滤后设置数据，即同一个公司（招商）
                        if row['源数据excel文件'] in file_name:       # 设置数据的源文件名字和当前文件名一致
                            print(row['源数据地址'], row['目标地址'], index_row)
                            try:
                                wb = load_workbook(file)
                                print("open file")
                                if str(row['源数据sheet表格']) in wb.sheetnames:
                                    sheet = wb[str(row['源数据sheet表格'])]
                                    src_value = sheet[str(row['源数据地址'])].value
                                    self.des_sheet[str(row['目标地址'])] = src_value
                                    print(src_value)
                                    self.text_edit.append(str(src_value))
                                else:
                                    print('源数据文件表格sheet错误')
                                    self.text_edit.append(self.error.format("源数据文件表格sheet错误"))
                            except FileNotFoundError as e:
                                print(f"源数据excel文件打开失败: {e}")
                                self.text_edit.append(self.error.format(f"源数据excel文件打开失败:{e}"))


    #遍历当前目录
    def list_directory(self):
        for entry in os.listdir(self.path):
            full_path = os.path.join(self.path, entry)

            if os.path.isdir(full_path):                    #如果是目录
                dir_name = os.path.basename(full_path)
                print("目录:"+dir_name)

                data = pd.DataFrame(self.set_data)          #根据目录过滤公司
                company_data = data[data['公司']==str(dir_name)]
                print(company_data)
                if company_data['源数据excel文件'].nunique() == 1 and company_data['源数据sheet表格'].nunique() == 1:
                    print("源文件和源sheet一致")
                    self.company_process(self, full_path, company_data)
                elif company_data['源数据excel文件'].nunique() != 1 :
                    print("源数据excel文件不一致")
                    self.text_edit.append(self.error.format("目录:" + str(dir_name)))
                    self.text_edit.append(self.error.format("源数据excel文件不一致"))
                elif company_data['源数据sheet表格'].nunique() != 1:
                    print("源sheet不一致")
                    self.text_edit.append(self.error.format("目录:" + str(dir_name)))
                    self.text_edit.append(self.error.format("源sheet不一致"))

    def start_button(self):
        print("开始")
        save_res = False
        temp_path = QtCore.QDir(self.line_edit_path.text())
        path = temp_path.absolutePath()
        print('目录:'+path)
        self.text_edit.clear()
        self.text_edit.append('目录:'+path)
        if not os.path.isdir(path):
            print("错误")
            self.text_edit.append(self.error.format("错误"))
        else:
            print("正确")
            self.text_edit.append(self.valid.format("目录正常"))

            subset_des_file_name = str(self.set_data.loc[0, '目标excel文件'])

            for index in os.listdir(path):  # 遍历该目录下文件
                des_file_path = os.path.join(path, index)
                des_file_name = os.path.basename(des_file_path)
                if subset_des_file_name in des_file_name:
                    des_file = Path(des_file_path)
                    if des_file.is_file():
                        try:
                            #获取目标sheet
                            parts = des_file_name.split('【')
                            parts = parts[1].split('】')
                            date = parts[0]
                            sheet_name = date[:4]+'.'+ date[4:]
                            print("目标文件:" + des_file_name)
                            print("目标sheet:" + sheet_name)

                            des_wb = load_workbook(des_file)
                            if str(sheet_name) in des_wb.sheetnames:
                                self.des_sheet = des_wb[str(sheet_name)]
                                self.text_edit.append(self.valid.format("目标文件打开正常"))
                                self.path = path
                                self.list_directory()
                                des_wb.save(des_file)
                                save_res = True
                                break
                            else:
                                print("未找到目标sheet:"+sheet_name)
                                self.text_edit.append(self.error.format("目标sheet打开失败:"+str(sheet_name)))
                                self.text_edit.append(self.error.format(str(des_wb.sheetnames)))

                        except FileNotFoundError as e:
                            save_res = False
                            print(f"目标文件打开失败 :{e}")
                            self.text_edit.append(self.error.format("目标文件"+des_file_name+"打开失败"))
                            self.text_edit.append(self.error.format(f"{e}"))
                else:
                    save_res = False
                    print("未找到目标文件")
                    self.text_edit.append(self.error.format("未找到目标文件"))

            if save_res:
                self.text_edit.append(self.valid.format("执行完成"))
            else:
                self.text_edit.append(self.error.format("执行失败"))

if __name__ == '__main__':
    # 每一个pyqt程序中都需要有一个QApplication对象，sys.argv是一个命令行参数列表
    app = QApplication(sys.argv)
    # 实例化主窗口并显示
    form = MainWindow()
    form.show()

    # 进入程序的主循环，遇到退出情况，终止程序
    sys.exit(app.exec_())