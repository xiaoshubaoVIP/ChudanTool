import os
from pathlib import Path

import pandas as pd
from PyQt5 import QtCore
from PyQt5.QtWidgets import QWidget, QPushButton, QLineEdit, QTextEdit, QVBoxLayout, QHBoxLayout, QFileDialog
from openpyxl import load_workbook


class DataAnalysis(QWidget):
    def __init__(self):
        super(DataAnalysis, self).__init__()
        self.error = '<font color="red">{}</font>'
        self.warning = '<font color="orange">{}</font>'
        self.valid = '<font color="green">{}</font>'

        #目标写入文件
        self.des_sheet = None

        #实例化按键
        self.btn = QPushButton("打开需要统计的目录")
        self.btn.setStyleSheet("background-color: rgb(255,255,255); color: black;")
        self.btn.setFixedHeight(40)
        self.btn.clicked.connect(self.get_dir)

        #文本框设定
        self.path = QtCore.QDir.currentPath()
        self.line_edit_path = QLineEdit(str(self.path))
        self.line_edit_path.setFixedHeight(40)
        self.line_edit_path.setStyleSheet("QLineEdit { background-color: white; }")

        #开始按键
        self.push_button_start = QPushButton('开始')
        self.push_button_start.setStyleSheet("background-color: rgb(255,255,255); color: black;")
        self.push_button_start.setFixedSize(100, 40)
        self.push_button_start.clicked.connect(self.start_button)

        #实例化textEdit并加入布局
        self.text_edit = QTextEdit()
        self.text_edit.setStyleSheet("QTextEdit { background-color: white; }")

        # 读设置信息
        temp_path = QtCore.QDir(self.path)
        self.set_path = temp_path.absolutePath()+'/setting/'
        if not os.path.isdir(self.set_path):
            os.mkdir(self.set_path)
            print("set文件不存在")
            self.text_edit.append(self.error.format("请先在setting目录下添加设置文件❌"))
        else:
            set_file = Path(self.set_path + 'setting.xlsx')
            if set_file.is_file():
                try:
                    self.set_data = pd.read_excel(set_file, dtype=str)  # 以字符串形式打开并读取excel表格
                    self.set_data = pd.DataFrame(self.set_data)
                    print(self.set_data)
                    print('set文件存在')
                    self.text_edit.append(self.valid.format("设置文件读取正常"))
                except FileNotFoundError as e:
                    print(f"设置文件打开失败: {e}")
                    self.text_edit.append(self.error.format(f"设置文件打开失败: {e}❌"))
            else:
                print("set文件不存在")
                self.text_edit.append(self.error.format("设置文件不存在❌"))

        #布局
        stack_main_layout = QVBoxLayout()
        stack_layout_1 = QHBoxLayout()
        stack_layout_2 = QHBoxLayout()

        #并加入布局
        stack_layout_1.addWidget(self.btn)
        stack_layout_1.addWidget(self.line_edit_path)
        stack_layout_1.addWidget(self.push_button_start)
        stack_layout_2.addWidget(self.text_edit)
        stack_main_layout.addLayout(stack_layout_1)
        stack_main_layout.addLayout(stack_layout_2)
        self.setLayout(stack_main_layout)

    @staticmethod
    def company_process(self, full_path, company_data):
        paths = os.walk(full_path)
        for path, dir_lst, file_lst in paths:
            for file_name in file_lst:
                file_path = os.path.join(path, file_name)
                file_name = os.path.basename(file_path)
                file = Path(str(file_path))
                print("文件:" + file_name)

                if file.is_file():
                    for index_row, row in company_data.iterrows():  # 遍历过滤后设置数据，即同一个公司（招商）
                        if row['源数据excel文件'] in file_name:       # 设置数据的源文件名字和当前文件名一致
                            print(row['源数据地址'], row['目标地址'], index_row)
                            try:
                                wb = load_workbook(file)
                                print("open file")
                                if str(row['源数据sheet表格']) in wb.sheetnames:
                                    sheet = wb[str(row['源数据sheet表格'])]
                                    src_value = sheet[str(row['源数据地址'])].value
                                    self.des_sheet[str(row['目标地址'])] = src_value
                                    print(src_value)
                                    self.text_edit.append(str(src_value))
                                else:
                                    print('源数据文件表格sheet错误')
                                    self.text_edit.append(self.error.format("源数据文件表格sheet错误❌"))
                            except FileNotFoundError as e:
                                print(f"源数据excel文件打开失败: {e}")
                                self.text_edit.append(self.error.format(f"源数据excel文件打开失败:{e}❌"))


    #遍历当前目录
    def list_directory(self):
        for entry in os.listdir(self.path):
            full_path = os.path.join(self.path, entry)

            if os.path.isdir(full_path):                    #如果是目录
                dir_name = os.path.basename(full_path)
                print("目录:"+dir_name)

                data = pd.DataFrame(self.set_data)          #根据目录过滤公司
                company_data = data[data['公司']==str(dir_name)]
                print(company_data)
                if company_data['源数据excel文件'].nunique() == 1 and company_data['源数据sheet表格'].nunique() == 1:
                    print("源文件和源sheet一致")
                    self.company_process(self, full_path, company_data)
                elif company_data['源数据excel文件'].nunique() != 1 :
                    print("源数据excel文件不一致")
                    self.text_edit.append(self.error.format("目录:" + str(dir_name)))
                    self.text_edit.append(self.error.format("源数据excel文件不一致❌"))
                elif company_data['源数据sheet表格'].nunique() != 1:
                    print("源sheet不一致")
                    self.text_edit.append(self.error.format("目录:" + str(dir_name)))
                    self.text_edit.append(self.error.format("源sheet不一致❌"))

    def get_dir(self):
        dialog = QFileDialog()
        dialog.options = QFileDialog.Options()
        dialog.options |= QFileDialog.ShowDirsOnly
        folder_path = QFileDialog.getExistingDirectory(self, "选择文件夹", options=dialog.options)

        if folder_path:
            print(f"选择的文件夹：{folder_path}")
            self.path = folder_path
            self.line_edit_path.setText(folder_path)

    def start_button(self):
        print("开始")
        save_res = False
        temp_path = QtCore.QDir(self.line_edit_path.text())
        path = temp_path.absolutePath()
        print('目录:'+path)
        self.text_edit.clear()
        self.text_edit.append('目录:'+path)
        if not os.path.isdir(path):
            print("错误")
            self.text_edit.append(self.error.format("目录错误❌"))
        else:
            print("正确")
            self.text_edit.append(self.valid.format("目录正常✅"))

            subset_des_file_name = str(self.set_data.loc[0, '目标excel文件'])

            for index in os.listdir(path):  # 遍历该目录下文件
                des_file_path = os.path.join(path, index)
                des_file_name = os.path.basename(des_file_path)
                if subset_des_file_name in des_file_name:
                    des_file = Path(des_file_path)
                    if des_file.is_file():
                        try:
                            #获取目标sheet
                            parts = des_file_name.split('【')
                            parts = parts[1].split('】')
                            date = parts[0]
                            sheet_name = date[:4]+'.'+ date[4:]
                            print("目标文件:" + des_file_name)
                            print("目标sheet:" + sheet_name)

                            des_wb = load_workbook(des_file)
                            if str(sheet_name) in des_wb.sheetnames:
                                self.des_sheet = des_wb[str(sheet_name)]
                                self.text_edit.append(self.valid.format("目标文件打开正常✅"))
                                self.path = path
                                self.list_directory()
                                des_wb.save(des_file)
                                save_res = True
                                break
                            else:
                                print("未找到目标sheet:"+sheet_name)
                                self.text_edit.append(self.error.format("目标sheet打开失败:"+str(sheet_name)+'❌'))
                                self.text_edit.append(self.error.format(str(des_wb.sheetnames)))

                        except FileNotFoundError as e:
                            save_res = False
                            print(f"目标文件打开失败 :{e}")
                            self.text_edit.append(self.error.format("目标文件"+des_file_name+"打开失败❌"))
                            self.text_edit.append(self.error.format(f"{e}"))
                else:
                    save_res = False
                    print("未找到目标文件")
                    self.text_edit.append(self.error.format("未找到目标文件❌"))

            if save_res:
                self.text_edit.append(self.valid.format("执行完成✅"))
            else:
                self.text_edit.append(self.error.format("执行失败❌"))